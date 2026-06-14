"""
将 Excel 打卡表转换为 history.json。

支持两种 Excel 格式：
  - 旧格式（第1周）：每周 7 天，每天 4 列（体重/运动/饮食/心得）
  - 新格式（第2周+）：每周 N 天，每天 7 列（体重/运动/饮食/心得/喊话/图片/评分）

用法：
  # 全新生成（第1周）
  python convert.py week1.xlsx

  # 追加第2周（从14天 Excel 中提取第8-14天）
  python convert.py week2data.xlsx --week 2 --start-day 8 --end-day 14 --append src/data/history.json
"""

import json
import sys
import os
from openpyxl import load_workbook

# 打卡人名称 → uid 直接映射（优先于 name_mapping）
NICKNAME_TO_UID = {
    "Latrell": "Latrell",
    "猪事顺利": "猪事顺利",
    "噤.": "噤.",
    "起个名字": "起个名字",
    "微信用户15f70e": "猪事顺利",
    "I miss": "I miss",
}

VALID_UIDS = {"Latrell", "Bard", "猪事顺利", "噤.", "起个名字", "I miss", "定轴转动的屑刚体"}


def load_name_mapping():
    config_path = os.path.join(os.path.dirname(__file__), 'config', 'name_mapping.json')
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def resolve_uid(name, name_mapping):
    """将任意名称解析为 uid。"""
    name = str(name).strip()
    if name in NICKNAME_TO_UID:
        return NICKNAME_TO_UID[name]
    if name in name_mapping:
        return name_mapping[name]
    if name in VALID_UIDS:
        return name
    print(f"  ⚠️  无法解析名称: '{name}'")
    return name


def detect_format(ws_content):
    """
    检测 Excel 格式。
    返回 (cols_per_day, total_days) — 每天几列，总共几天。
    """
    # 旧格式：row1 每天1个日期 + 5个空 → 6列/天 → 每天4列数据
    # 新格式：row1 每天1个日期 + 6个空 → 7列/天 → 每天7列数据
    total_days = 0
    for col in range(5, ws_content.max_column + 1):
        val = ws_content.cell(row=1, column=col).value
        if val and str(val).startswith('2026'):
            total_days += 1

    # 列数推算：旧格式日期间隔为6（4字段+2空），新格式为7（7字段）
    # 更简单的检测：看 row2 的 col+4 是什么
    # 旧格式: col+4 = None (因为是空列或下一组日期)
    # 新格式: col+4 = "隔空喊话"
    first_date_col = None
    for col in range(5, ws_content.max_column + 1):
        val = ws_content.cell(row=1, column=col).value
        if val and str(val).startswith('2026'):
            first_date_col = col
            break

    if first_date_col is None:
        return 4, total_days  # fallback to old format

    # 旧格式: 间隔为6（每天6列 = 4数据 + 2空位）
    # 新格式: 间隔为7（每天7列）
    next_date_col = None
    for col in range(first_date_col + 1, ws_content.max_column + 1):
        val = ws_content.cell(row=1, column=col).value
        if val and str(val).startswith('2026'):
            next_date_col = col
            break

    cols_per_day = (next_date_col - first_date_col) if next_date_col else 7
    return cols_per_day, total_days


def parse_excel_content(excel_path, name_mapping, start_day, end_day, cols_per_day):
    """
    解析 汇总-内容 sheet，返回 participants_data。
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['汇总-内容']

    # 计算有效范围
    total_days = 0
    date_col_map = {}  # day_idx → (date_str, base_col)
    for col in range(5, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).startswith('2026'):
            total_days += 1
            date_col_map[total_days] = (str(val).strip(), col)

    print(f"   检测到 {total_days} 天数据，每天 {cols_per_day} 列")

    # 确定提取范围（1-based day index）
    first_day = start_day if start_day else 1
    last_day = end_day if end_day else total_days
    print(f"   提取第 {first_day}-{last_day} 天")

    participants_data = []
    for row_idx in range(3, 10):
        raw_name = str(ws.cell(row=row_idx, column=2).value or '')
        uid = resolve_uid(raw_name, name_mapping)

        daily_records = []
        for day_idx in range(first_day, last_day + 1):
            date_str, base_col = date_col_map[day_idx]

            weight_raw = ws.cell(row=row_idx, column=base_col).value
            sport = str(ws.cell(row=row_idx, column=base_col + 1).value or '')
            diet = str(ws.cell(row=row_idx, column=base_col + 2).value or '')
            note = str(ws.cell(row=row_idx, column=base_col + 3).value or '')

            # 新格式有 shoutOut（隔空喊话），旧格式没有
            shout_out = ''
            if cols_per_day >= 7:
                shout_out = str(ws.cell(row=row_idx, column=base_col + 4).value or '')

            # 检查是否打卡（有体重数据即为打卡）
            has_checkin = (
                weight_raw is not None
                and str(weight_raw).strip() != ''
                and str(weight_raw).strip() != 'None'
            )

            if has_checkin:
                try:
                    weight = float(weight_raw)
                except (ValueError, TypeError):
                    weight = None
            else:
                weight = None

            if sport in ('None', ''):
                sport = '未运动' if not has_checkin else ''
            if diet in ('None', ''):
                diet = '未控制'
            if note in ('None', ''):
                note = ''
            if shout_out in ('None', ''):
                shout_out = ''

            record = {
                "date": date_str,
                "weight": weight,
                "sport": sport,
                "diet": diet,
                "note": note,
            }
            if cols_per_day >= 7:
                record["shoutOut"] = shout_out

            daily_records.append(record)

        participants_data.append({
            "uid": uid,
            "dailyRecords": daily_records,
            "aiComment": {
                "uid": uid,
                "title": "",
                "tags": [],
                "highlight": "",
                "comment": "",
                "nextWeekFlag": "",
                "prediction": "",
                "coachGuide": "",
                "sassQuote": "",
                "sassReply": "",
            },
        })

    return participants_data


def build_week_data(excel_path, week_num, start_day, end_day, name_mapping):
    """构建单周 WeekData。"""
    wb = load_workbook(excel_path, data_only=True)
    ws_content = wb['汇总-内容']
    cols_per_day, total_days = detect_format(ws_content)

    participants_data = parse_excel_content(
        excel_path, name_mapping, start_day, end_day, cols_per_day
    )

    # 计算日期范围
    dates = sorted(set(
        r['date'] for p in participants_data for r in p['dailyRecords']
    ))
    first_date = dates[0] if dates else ''
    last_date = dates[-1] if dates else ''
    date_range = f"{first_date} - {last_date}"

    return {
        "week": week_num,
        "dateRange": date_range,
        "participants": participants_data,
    }


def append_to_history(week_data, history_path):
    """将一周数据追加到现有 history.json。"""
    with open(history_path, 'r', encoding='utf-8') as f:
        history = json.load(f)

    existing_weeks = {w['week'] for w in history}
    if week_data['week'] in existing_weeks:
        print(f"  ⚠️  第 {week_data['week']} 周已存在，覆盖旧数据")

    history.append(week_data)
    # 按周号排序
    history.sort(key=lambda w: w['week'])

    with open(history_path, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

    print(f"✅ 已追加第 {week_data['week']} 周到 {history_path}")
    print(f"   日期: {week_data['dateRange']}")
    print(f"   共 {len(history)} 周，{len(week_data['participants'])} 名参赛者")


def main():
    args = sys.argv[1:]

    excel_path = None
    output_path = 'src/data/history.json'
    week_num = 1
    start_day = None
    end_day = None
    append_mode = False

    i = 0
    while i < len(args):
        arg = args[i]
        if arg == '--week' and i + 1 < len(args):
            i += 1
            week_num = int(args[i])
        elif arg == '--start-day' and i + 1 < len(args):
            i += 1
            start_day = int(args[i])
        elif arg == '--end-day' and i + 1 < len(args):
            i += 1
            end_day = int(args[i])
        elif arg == '--append':
            append_mode = True
        elif not arg.startswith('--') and not excel_path:
            excel_path = arg
        elif not arg.startswith('--'):
            output_path = arg
        i += 1

    if not excel_path:
        print("用法: python convert.py <excel_path> [--week N] [--start-day N] [--end-day N] [--append] [output_path]")
        sys.exit(1)

    name_mapping = load_name_mapping()
    week_data = build_week_data(excel_path, week_num, start_day, end_day, name_mapping)

    if append_mode:
        append_to_history(week_data, output_path)
    else:
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump([week_data], f, ensure_ascii=False, indent=2)
        print(f"✅ 已生成 {output_path}")
        print(f"   日期: {week_data['dateRange']}")
        print(f"   共 {len(week_data['participants'])} 名参赛者")


if __name__ == '__main__':
    main()
